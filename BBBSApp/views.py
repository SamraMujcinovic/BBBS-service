import os
from datetime import datetime
from io import BytesIO

import openpyxl
from django.core.mail import send_mail
from django.http import HttpResponse
from django.utils.encoding import force_str, force_bytes
from openpyxl.styles import NamedStyle, Border, Side
from rest_framework.status import HTTP_409_CONFLICT, HTTP_403_FORBIDDEN
from rest_framework_simplejwt.tokens import RefreshToken
from rest_framework_simplejwt.token_blacklist.models import (
    OutstandingToken,
    BlacklistedToken,
)

from rest_framework_simplejwt.views import (
    TokenRefreshView,
)

from rest_framework_simplejwt.authentication import JWTAuthentication

from django.middleware import csrf
from django.contrib.auth import authenticate
from django.conf import settings
from rest_framework import status
from django.contrib.auth.hashers import check_password

from django.utils.http import urlsafe_base64_decode, urlsafe_base64_encode
from django.contrib.auth.tokens import default_token_generator
from django.shortcuts import get_object_or_404

# import viewsets
from rest_framework import viewsets
from rest_framework.views import APIView

from rest_framework.response import Response

from rest_framework.permissions import AllowAny, BasePermission, IsAuthenticated

from .pagination import CustomPagination
# import local data
from .serializers import (
    ChildSerializer,
    CoordinatorSerializer,
    FormSerializer,
    VolunteerSerializer,
    VolunteerHoursSerializer,
    Organisation_Serializer,
    City_Serializer,
    MentoringReasonSerializer,
    Developmental_DifficultiesSerializer,
    MentoringReasonCategorySerializer,
    HangOutPlaceSerializer,
    ActivitiesSerializer,
    ActivityCategorySerializer,
    CustomTokenRefreshSerializer
)
from django.db.models import Sum, F
from .models import (
    Child,
    Coordinator,
    Coordinator_Organisation_City,
    Form,
    Volunteer,
    Organisation,
    City,
    Mentoring_Reason,
    Developmental_Difficulties,
    Mentoring_Reason_Category,
    Hang_Out_Place,
    Activities,
    Activity_Category,
    Volunteer_Organisation_City,
    Child_Organisation_City
)
from django.contrib.auth.models import User

from .utilis import isUserAdmin, isUserCoordinator, isUserVolunteer


class IsAdmin(BasePermission):
    def has_permission(self, request, view):
        return request.user.groups.filter(name="admin").exists()

    def has_object_permission(self, request, view, obj):
        return request.user.groups.filter(name="admin").exists()


class IsCoordinator(BasePermission):
    def has_permission(self, request, view):
        return request.user.groups.filter(name="coordinator").exists()

    def has_object_permission(self, request, view, obj):
        return request.user.groups.filter(name="coordinator").exists()


class IsVolunteer(BasePermission):
    def has_permission(self, request, view):
        return request.user.groups.filter(name="volunteer").exists()

    def has_object_permission(self, request, view, obj):
        return request.user.groups.filter(name="volunteer").exists()


class OrganisationView(viewsets.ModelViewSet):
    authentication_classes = [JWTAuthentication]
    pagination_class = CustomPagination
    def get_permissions(self):
        permission_classes = []
        if self.action == "create":
            permission_classes = [IsAdmin]
        elif self.action == "list":
            permission_classes = [IsCoordinator | IsAdmin | IsVolunteer]
        elif self.action == "delete":
            permission_classes = [IsAdmin]
        return [permission() for permission in permission_classes]

        # define queryset
    def get_queryset(self):
        user = self.request.user
        if isUserAdmin(user):
            return Organisation.objects.all().order_by('name')
        if isUserCoordinator(user):
            coordinator = Coordinator.objects.get(user_id=user.id)
            coordinator_organisation_city = Coordinator_Organisation_City.objects.get(
                coordinator_id=coordinator.id
            )
            return Organisation.objects.filter(id=coordinator_organisation_city.organisation_id)
        if isUserVolunteer(user):
            volunteer = Volunteer.objects.get(user_id=user.id)
            volunteer_organisation_city = Volunteer_Organisation_City.objects.get(
                volunteer_id=volunteer.id
            )
            return Organisation.objects.filter(id=volunteer_organisation_city.organisation_id)
        return None

    def destroy(self, request, *args, **kwargs):
        organisation = self.get_object()

        if checkIfOrganisationIsInUse(organisation.id):
            return Response({"This organisation is being used by a coordinator, volunteer or child and cannot be deleted."},status=HTTP_409_CONFLICT)

        # If no coordinators are associated, proceed with deletion
        return super().destroy(request, *args, **kwargs)

    serializer_class = Organisation_Serializer

def checkIfOrganisationIsInUse(organisation_id):
    return Coordinator_Organisation_City.objects.filter(organisation_id=organisation_id).exists() or Volunteer_Organisation_City.objects.filter(organisation_id=organisation_id).exists() or Child_Organisation_City.objects.filter(organisation_id=organisation_id).exists()


class CityView(viewsets.ModelViewSet):
    authentication_classes = [JWTAuthentication]
    queryset = City.objects.all()
    serializer_class = City_Serializer


class DevelopmentalDifficultiesView(viewsets.ModelViewSet):
    authentication_classes = [JWTAuthentication]
    queryset = Developmental_Difficulties.objects.all().order_by("id")
    serializer_class = Developmental_DifficultiesSerializer


class MentoringReasonView(viewsets.ModelViewSet):
    authentication_classes = [JWTAuthentication]
    queryset = Mentoring_Reason.objects.all().order_by("id")
    serializer_class = MentoringReasonSerializer


class MentoringReasonCategoryView(viewsets.ModelViewSet):
    authentication_classes = [JWTAuthentication]
    queryset = Mentoring_Reason_Category.objects.all().order_by("id")
    serializer_class = MentoringReasonCategorySerializer


class HangOutPlaceView(viewsets.ModelViewSet):
    authentication_classes = [JWTAuthentication]
    queryset = Hang_Out_Place.objects.all().order_by("id")
    serializer_class = HangOutPlaceSerializer


class ActivitiesView(viewsets.ModelViewSet):
    authentication_classes = [JWTAuthentication]
    queryset = Activities.objects.all()
    serializer_class = ActivitiesSerializer


class ActivityCategoryView(viewsets.ModelViewSet):
    authentication_classes = [JWTAuthentication]
    queryset = Activity_Category.objects.all()
    serializer_class = ActivityCategorySerializer


# create a viewset
class CoordinatorView(viewsets.ModelViewSet):
    authentication_classes = [JWTAuthentication]
    pagination_class = CustomPagination
    def get_permissions(self):
        permission_classes = []
        if self.action == "create":
            permission_classes = [IsAdmin]
        elif self.action == "list":
            permission_classes = [IsCoordinator | IsAdmin]
        elif self.action == "destroy":
            permission_classes = [IsAdmin]
        return [permission() for permission in permission_classes]

    # define queryset
    def get_queryset(self):
        user = self.request.user
        resultset = []
        if isUserAdmin(user):
            resultset = Coordinator.objects.all().order_by('coordinator_organisation', 'coordinator_city', 'user__first_name', 'user__last_name')
            if self.request.GET.get("organisation") is not None:
                coordinator_organisation = self.request.GET.get("organisation")
                resultset = resultset.filter(coordinator_organisation=coordinator_organisation)
            if self.request.GET.get("city") is not None:
                coordinator_city = self.request.GET.get("city")
                resultset = resultset.filter(coordinator_city=coordinator_city)
            return resultset
        if isUserCoordinator(user):
            return Coordinator.objects.filter(user_id=user.id)
        return None

    def destroy(self, request, *args, **kwargs):
        current_user = self.request.user
        coordinator = self.get_object()

        if isUserCoordinator(current_user) or isUserVolunteer(current_user):
            return Response(status=HTTP_403_FORBIDDEN)
        if checkIfCoordinatorIsInUse(coordinator.id):
            return Response({"This coordinator is being used by a volunteer or child and cannot be deleted."},status=HTTP_409_CONFLICT)

        return super().destroy(request, *args, **kwargs)

    # specify serializer to be used
    serializer_class = CoordinatorSerializer


def checkIfCoordinatorIsInUse(coordinator_id):
    return Volunteer.objects.filter(coordinator__id=coordinator_id).exists() or Child.objects.filter(coordinator__id=coordinator_id).exists()


# create a viewset
class VolunteerView(viewsets.ModelViewSet):
    authentication_classes = [JWTAuthentication]
    pagination_class = CustomPagination
    # define queryset
    def get_queryset(self):
        user = self.request.user
        return get_accessible_volunteers(user, self.request.GET)

    def get_permissions(self):
        permission_classes = []
        if self.action == "create":
            permission_classes = [IsAdmin | IsCoordinator]
        elif self.action == "list":
            permission_classes = [IsAdmin | IsCoordinator | IsVolunteer]
        elif self.action == "destroy":
            permission_classes = [IsAdmin | IsCoordinator]
        return [permission() for permission in permission_classes]

    def destroy(self, request, *args, **kwargs):
        current_user = self.request.user
        volunteer = self.get_object()

        if isUserVolunteer(current_user):
            return Response(status=HTTP_403_FORBIDDEN)
        if isUserCoordinator(current_user):
            # forbid coordinators to delete forms that are not in his organisation
            coordinator = Coordinator.objects.filter(user_id=current_user.id).first()
            if volunteer.volunteer_organisation.first().id != coordinator.coordinator_organisation.first().id or volunteer.volunteer_city.first().id != coordinator.coordinator_city.first().id:
                return Response(
                    {"error": "Coordinator is not allowed to delete volunteers from other organisations!"},
                    status=status.HTTP_403_FORBIDDEN
                )
        if checkIfVolunteerIsInUse(volunteer.id):
            return Response({"This volunteer is being used by a child or form and cannot be deleted."},status=HTTP_409_CONFLICT)

        return super().destroy(request, *args, **kwargs)

    # specify serializer to be used
    serializer_class = VolunteerSerializer


def get_accessible_volunteers(user, filters):
    resultset = []
    if isUserAdmin(user):
        resultset = Volunteer.objects.all()
    if isUserCoordinator(user):
        # allow coordinators to see volunteers from his organisation and city
        coordinator = Coordinator.objects.get(user_id=user.id)
        coordinator_organisation_city = Coordinator_Organisation_City.objects.get(
            coordinator_id=coordinator.id
        )
        resultset = Volunteer.objects.filter(
            volunteer_organisation=coordinator_organisation_city.organisation_id,
            volunteer_city=coordinator_organisation_city.city_id
        ).order_by('user__first_name', 'user__last_name')
    if isUserVolunteer(user):
        resultset = Volunteer.objects.filter(user_id=user.id)

    # when organisation filter is selected, get data from thar org only
    if filters.get("organisationFilter") is not None:
        organisation = filters.get("organisationFilter")
        return resultset.filter(volunteer_organisation=organisation)

    if filters.get("status") is not None:
        volunteer_status = filters.get("status")
        resultset = resultset.filter(status=volunteer_status)
    if filters.get("gender") is not None:
        volunteer_gender = filters.get("gender")
        resultset = resultset.filter(gender=volunteer_gender)
    if filters.get("organisation") is not None:
        volunteer_organisation = filters.get("organisation")
        resultset = resultset.filter(volunteer_organisation=volunteer_organisation)
    if filters.get("city") is not None:
        volunteer_city = filters.get("city")
        resultset = resultset.filter(volunteer_city=volunteer_city)
    if (filters.get("availableVolunteers") is not None and
            filters.get("availableVolunteers") == "True"):
        resultset = resultset.filter(child=None)
    if filters.get("child") is not None:
        child_volunteer = Volunteer.objects.filter(child=filters.get("child"))
        # add current childs volunteer to list of accessible volunteers
        if child_volunteer.first() is not None:
            resultset = resultset | child_volunteer

    return resultset.order_by('volunteer_organisation', 'volunteer_city', 'user__first_name', 'user__last_name')


def checkIfVolunteerIsInUse(volunteer_id):
    return Child.objects.filter(volunteer__id=volunteer_id).exists() or Form.objects.filter(volunteer__id=volunteer_id).exists()


# create a viewset
class ChildView(viewsets.ModelViewSet):
    authentication_classes = [JWTAuthentication]
    pagination_class = CustomPagination
    # define queryset
    def get_queryset(self):
        user = self.request.user
        resultset = []
        if isUserAdmin(user):
            resultset = Child.objects.all()
        if isUserCoordinator(user):
            # allow coordinators to see childs from his organisation and city
            coordinator = Coordinator.objects.get(user_id=user.id)
            coordinator_organisation_city = Coordinator_Organisation_City.objects.get(
                coordinator_id=coordinator.id
            )
            resultset = Child.objects.filter(
                child_organisation=coordinator_organisation_city.organisation_id,
                child_city=coordinator_organisation_city.city_id
            )

        # when organisation filter is selected, get data from that org only
        if self.request.GET.get("organisationFilter") is not None:
            organisation = self.request.GET.get("organisationFilter")
            return resultset.filter(child_organisation=organisation)

        if self.request.GET.get("organisation") is not None:
            child_organisation = self.request.GET.get("organisation")
            resultset = resultset.filter(child_organisation=child_organisation)
        if self.request.GET.get("city") is not None:
            child_city = self.request.GET.get("city")
            resultset = resultset.filter(child_city=child_city)
        return resultset.order_by('child_organisation', 'child_city', 'code')

    def get_permissions(self):
        permission_classes = []
        if self.action == "create":
            permission_classes = [IsAdmin | IsCoordinator]
        elif self.action == "list":
            permission_classes = [IsAdmin | IsCoordinator]
        elif self.action == "destroy":
            permission_classes = [IsAdmin | IsCoordinator]
        return [permission() for permission in permission_classes]

    def destroy(self, request, *args, **kwargs):
        current_user = self.request.user
        child = self.get_object()

        if isUserVolunteer(current_user):
            return Response(status=HTTP_403_FORBIDDEN)
        if isUserCoordinator(current_user):
            # forbid coordinators to delete child that are not in his organisation
            coordinator = Coordinator.objects.filter(user_id=current_user.id).first()
            if child.child_organisation.first().id != coordinator.coordinator_organisation.first().id or child.child_city.first().id != coordinator.coordinator_city.first().id:
                return Response(
                    {"error": "Coordinator is not allowed to delete child from other organisations!"},
                    status=status.HTTP_403_FORBIDDEN
                )
        if checkIfChildIsInUse(child.id):
            return Response({"This child is being used by a volunteer and cannot be deleted."},status=HTTP_409_CONFLICT)


        return super().destroy(request, *args, **kwargs)

    # specify serializer to be used
    serializer_class = ChildSerializer


def checkIfChildIsInUse(child_id):
    return Volunteer.objects.filter(child__id=child_id).exists()

# create a viewset
class FormView(viewsets.ModelViewSet):
    authentication_classes = [JWTAuthentication]
    pagination_class = CustomPagination
    # define queryset
    def get_queryset(self):
        return filter_accessible_forms(self.request.user, self.request.GET)

    def get_permissions(self):
        permission_classes = []
        if self.action == "create":
            permission_classes = [IsVolunteer]
        elif self.action == "list":
            permission_classes = [IsAdmin | IsCoordinator | IsVolunteer]
        elif self.action == "destroy":
            permission_classes = [IsAdmin | IsCoordinator]
        return [permission() for permission in permission_classes]

    def destroy(self, request, *args, **kwargs):
        current_user = self.request.user
        form = self.get_object()

        if isUserVolunteer(current_user):
            return Response(status=HTTP_403_FORBIDDEN)
        if isUserCoordinator(current_user):
            # forbid cooridnators to delete forms that are not in his organisation
            accessible_forms_ids = set(get_accessible_forms(self.request.user).values_list('id', flat=True))
            if form.id not in accessible_forms_ids:
                return Response(
                    {"error": "Coordinator is not allowed to delete forms from other organisations!"},
                    status=status.HTTP_403_FORBIDDEN
                )

        return super().destroy(request, *args, **kwargs)

    # specify serializer to be used
    serializer_class = FormSerializer


def get_accessible_forms(current_user):
    if isUserAdmin(current_user):
        return Form.objects.all().order_by("-date", 'volunteer__volunteer_organisation', 'volunteer__volunteer_city')
    if isUserCoordinator(current_user):
        # allow coordinators to see forms of his volunteer
        coordinator = Coordinator.objects.get(user_id=current_user.id)
        return Form.objects.filter(
            volunteer__volunteer_organisation__in=coordinator.coordinator_organisation.all(),
            volunteer__volunteer_city__in=coordinator.coordinator_city.all()
        ).order_by("-date")
    if isUserVolunteer(current_user):
        # allow volunteers to see his forms
        volunteer = Volunteer.objects.get(user_id=current_user.id)
        return Form.objects.filter(volunteer=volunteer.id).order_by("-date")
    return None


def filter_accessible_forms(current_user, filters):
    resultset = get_accessible_forms(current_user)
    if filters.get("organisationFilter") is not None:
        organisation = filters.get("organisationFilter")
        resultset = resultset.filter(volunteer__volunteer_organisation__id=organisation)
        if not resultset:
            return resultset
    if filters.get("volunteerFilter") is not None:
        volunteer = filters.get("volunteerFilter")
        resultset = resultset.filter(volunteer_id=volunteer)
        if not resultset:
            return resultset
    if filters.get("activityTypeFilter") is not None:
        activity_type = filters.get("activityTypeFilter")
        resultset = resultset.filter(activity_type=Form.ACTIVITY_TYPE_DICT.get(activity_type))
        if not resultset:
            return resultset
    if filters.get("startDate") is not None and filters.get("endDate") is not None:
        start_date = filters.get("startDate")
        end_date = filters.get("endDate")
        resultset = resultset.filter(date__range=[start_date, end_date])
        if not resultset:
            return resultset

    return  resultset


class VolunteerHours(viewsets.ModelViewSet):
    authentication_classes = [JWTAuthentication]
    pagination_class = CustomPagination
    # define queryset
    def get_queryset(self):
        user = self.request.user
        start_date = self.request.GET.get("startDate")
        end_date = self.request.GET.get("endDate")
        if isUserAdmin(user):
            return (Form
                    .objects
                    .filter(date__range=[start_date, end_date])
                    .values(
                        volunteer_user_id=F('volunteer__user_id'),
                        volunteer_first_name=F('volunteer__user__first_name'),
                        volunteer_last_name=F('volunteer__user__last_name'),
                        volunteer_organisation=F('volunteer__volunteer_organisation__name'),
                        volunteer_city=F('volunteer__volunteer_city__name')
                    )
                    .annotate(volunteer_hours=Sum('duration'))
                    .order_by('volunteer__user__first_name'))
        elif isUserCoordinator(user):
            coordinator = Coordinator.objects.get(user_id=user.id)
            return (Form
                    .objects
                    .filter(volunteer__volunteer_organisation__in=coordinator.coordinator_organisation.all(),
                            volunteer__volunteer_city__in=coordinator.coordinator_city.all(),
                            date__range=[start_date, end_date])
                    .values(
                            volunteer_user_id=F('volunteer__user_id'),
                            volunteer_first_name=F('volunteer__user__first_name'),
                            volunteer_last_name=F('volunteer__user__last_name'),
                            volunteer_organisation=F('volunteer__volunteer_organisation__name'),
                            volunteer_city=F('volunteer__volunteer_city__name')
                    )
                    .annotate(volunteer_hours=Sum('duration'))
                    .order_by('volunteer__user__first_name'))
        elif isUserVolunteer(user):
            volunteer = Volunteer.objects.get(user_id=user.id)
            return (Form
                    .objects
                    .filter(volunteer=volunteer.id, date__range=[start_date, end_date])
                    .values(
                            volunteer_user_id=F('volunteer__user_id'),
                            volunteer_first_name=F('volunteer__user__first_name'),
                            volunteer_last_name=F('volunteer__user__last_name'),
                            volunteer_organisation=F('volunteer__volunteer_organisation__name'),
                            volunteer_city=F('volunteer__volunteer_city__name')
                    )
                    .annotate(volunteer_hours=Sum('duration'))
                    .order_by('volunteer__user__first_name'))
        return None

    def get_permissions(self):
        permission_classes = []
        if self.action == "list":
            permission_classes = [IsAdmin | IsCoordinator | IsVolunteer]
        return [permission() for permission in permission_classes]

    # specify serializer to be used
    serializer_class = VolunteerHoursSerializer


class FormsTotalHoursSumView(APIView):
    authentication_classes = [JWTAuthentication]
    permission_classes = (AllowAny,)

    def get(self, request):
        resultset = get_accessible_forms(self.request.user)
        if self.request.GET.get("organisationFilter") is not None:
            organisation = self.request.GET.get("organisationFilter")
            resultset = resultset.filter(volunteer__volunteer_organisation__id=organisation)
        if self.request.GET.get("volunteerFilter") is not None:
            volunteer = self.request.GET.get("volunteerFilter")
            resultset = resultset.filter(volunteer_id=volunteer)
        if self.request.GET.get("activityTypeFilter") is not None:
            activity_type = self.request.GET.get("activityTypeFilter")
            resultset = resultset.filter(activity_type=Form.ACTIVITY_TYPE_DICT.get(activity_type))
        if self.request.GET.get("startDate") is not None and self.request.GET.get("endDate") is not None:
            start_date = self.request.GET.get("startDate")
            end_date = self.request.GET.get("endDate")
            resultset = resultset.filter(date__range=[start_date, end_date])

        return None if resultset is None else Response(data=resultset.aggregate(totalHours=Sum('duration')), status=status.HTTP_200_OK)


def get_tokens_for_user(user):
    refresh = RefreshToken.for_user(user)
    refresh["username"] = user.username
    refresh["first_name"] = user.first_name
    refresh["last_name"] = user.last_name
    refresh["group"] = list(user.groups.values_list('name',flat=True))
    return {
        'refresh': str(refresh),
        'access': str(refresh.access_token),
    }


class LoginView(APIView):
    permission_classes = (AllowAny,)
    def post(self, request, format=None):
        data = request.data
        response = Response()
        username = data.get('username', None)
        password = data.get('password', None)
        user = authenticate(username=username, password=password)

        if user is not None:
            if user.is_active:
                data = get_tokens_for_user(user)
                response.set_cookie(
                    key=settings.SIMPLE_JWT['AUTH_COOKIE'],
                    value=data["refresh"],
                    expires=settings.SIMPLE_JWT['REFRESH_TOKEN_LIFETIME'],
                    httponly=settings.SIMPLE_JWT['AUTH_COOKIE_HTTP_ONLY'],
                    samesite=settings.SIMPLE_JWT['AUTH_COOKIE_SAMESITE'],
                    secure=settings.SIMPLE_JWT['AUTH_COOKIE_SECURE']
                )
                csrf.get_token(request)
                response.data = {"Success": "Login successfully", "data": data["access"]}
                return response
            else:
                return Response({"No active": "This account is not active!!"}, status=status.HTTP_404_NOT_FOUND)
        else:
            return Response({"Invalid": "Invalid username or password!!"}, status=status.HTTP_404_NOT_FOUND)


class CustomTokenRefreshView(TokenRefreshView):
    permission_classes = [AllowAny]
    serializer_class = CustomTokenRefreshSerializer


class LogoutView(APIView):
    authentication_classes = [JWTAuthentication,]
    permission_classes = (AllowAny,)

    def post(self, request, *args, **kwargs):
        if self.request.data.get("all"):
            token: OutstandingToken
            for token in OutstandingToken.objects.filter(user=request.user):
                _, _ = BlacklistedToken.objects.get_or_create(token=token)
            return Response({"status": "OK, goodbye, all refresh tokens blacklisted"})
        refresh_token = self.request.data.get("refresh_token")
        token = RefreshToken(token=refresh_token)
        token.blacklist()
        return Response({"status": "OK, goodbye"})


class PasswordChangeView(APIView):
    authentication_classes = [JWTAuthentication]
    permission_classes = (IsAuthenticated,)

    def post(self, request):
        print(request.data.get("newPassword", None))
        if request.data.get("newPassword", None) is None or request.data.get("oldPassword", None) is None:
            return Response(status=status.HTTP_400_BAD_REQUEST)

        current_user = request.user

        if not check_password(request.data.get("oldPassword"), current_user.password):
            return Response({"oldPassword": "Invalid!"}, status=status.HTTP_400_BAD_REQUEST)

        current_user.set_password(request.data.get("newPassword"))
        current_user.save()

        return Response(status=status.HTTP_200_OK)


class RequestPasswordResetView(APIView):
    permission_classes = (AllowAny,)

    def post(self, request):
        email = request.data.get('email')
        if not email:
            return Response({'error': 'Email is required'}, status=status.HTTP_400_BAD_REQUEST)

        # Check if the user with the provided email exists
        try:
            user = User.objects.get(email=email)
        except User.DoesNotExist:
            return Response({'error': 'No user associated with this email'}, status=status.HTTP_404_NOT_FOUND)

        # Generate a reset token and URL-safe user ID
        token = default_token_generator.make_token(user)
        uid = urlsafe_base64_encode(force_bytes(user.pk))
        reset_link = f"{settings.CLIENT_URL}/reset-password/{uid}/{token}/"

        # Send the reset email
        send_mail(
            'Oporavak lozinke',
            f"Kliknite na link da unesete novu lozinku:\n{reset_link}",
            None,
            [user.email],
            fail_silently=False,
        )

        return Response({'success': 'Password reset email sent successfully'}, status=status.HTTP_200_OK)


class PasswordResetConfirmView(APIView):
    permission_classes = (AllowAny,)

    def post(self, request):
        uidb64 = request.data.get('uid')
        token = request.data.get('token')
        new_password = request.data.get('password')

        # Decode the user ID from the base64 string
        try:
            uid = force_str(urlsafe_base64_decode(uidb64))
            user = User.objects.get(pk=uid)
        except (TypeError, ValueError, OverflowError, User.DoesNotExist):
            return Response({'error': 'Invalid user ID'}, status=status.HTTP_400_BAD_REQUEST)

        # Validate the token
        if not default_token_generator.check_token(user, token):
            return Response({'error': 'Invalid or expired token'}, status=status.HTTP_400_BAD_REQUEST)

        # Set the new password
        user.set_password(new_password)
        user.save()

        return Response({'success': 'Password has been reset successfully'}, status=status.HTTP_200_OK)


class EmailRemindersView(APIView):
    permission_classes = (AllowAny,)

    def post(self, request):
        start_date = datetime.strptime(request.data.get("start_date", None), '%Y-%m-%d')
        end_date = datetime.strptime(request.data.get("end_date", None), '%Y-%m-%d')

        if request.data.get("volunteer_user_id", None) is None or start_date is None or end_date is None:
            return Response(status=status.HTTP_400_BAD_REQUEST)

        user = User.objects.filter(id=request.data.get("volunteer_user_id")).first()
        if user is None:
            return Response(status=status.HTTP_404_NOT_FOUND)

        email_message = (
                "Poštovani,\n\nVaši sati za period od "
                + start_date.strftime("%d.%m.%Y")
                + " do "
                + end_date.strftime("%d.%m.%Y")
                + " nisu dovoljni. Minimalan broj sati u mjesecu je 16. Molimo Vas da unesete preostale sate.\n\nUnaprijed zahvaljujemo!"
        )

        send_mail(
            "Podsjetnik za unos sati",
            email_message,
            None,
            [user.email],
            fail_silently=False,
        )
        return Response(status=status.HTTP_200_OK)


class ActivateUser(APIView):
    permission_classes = (AllowAny,)

    def post(self, request):
        uid = request.data.get('uid')
        token = request.data.get('token')
        password = request.data.get('password')

        try:
            # Decode the user ID
            uid = force_str(urlsafe_base64_decode(uid))
            user = get_object_or_404(User, pk=uid)

            # Check the token validity
            if default_token_generator.check_token(user, token):
                # Set the new password
                user.set_password(password)
                user.is_active = True  # Activate the user
                user.save()

                return Response({'message': 'Account activated successfully.'}, status=status.HTTP_200_OK)
            else:
                return Response({'error': 'Invalid token or token expired.'}, status=status.HTTP_400_BAD_REQUEST)
        except (TypeError, ValueError, OverflowError, User.DoesNotExist):
            return Response({'error': 'Invalid activation link.'}, status=status.HTTP_400_BAD_REQUEST)


class FormsExcelView(APIView):
    authentication_classes = [JWTAuthentication]
    permission_classes = (IsAuthenticated,)

    def get(self, request):
        data = filter_accessible_forms(self.request.user, self.request.GET)
        return excel_file_generation(data, "forms", fill_rows_in_forms_excel, 10, getFormsExcelFileName(self.request.GET))



class VolunteersExcelView(APIView):
    authentication_classes = [JWTAuthentication]
    permission_classes = (IsAuthenticated,)

    def get(self, request):
        data = get_accessible_volunteers(self.request.user, self.request.GET)
        return excel_file_generation(data, "volunteers", fill_rows_in_volunteers_excel, 14, getVolunteersExcelFileName(self.request.GET))


def excel_file_generation(data, template_name, function, number_of_columns, file_name):
    # Load the predefined Excel template
    template_path = os.path.join(os.path.dirname(__file__), "templates",
                                 f"{template_name}.xlsx")  # Update with the path to your template
    try:
        workbook = openpyxl.load_workbook(template_path)
    except Exception as e:
        print(f"Error loading template: {e}")
        raise

    # Access the desired sheet in the template
    sheet = workbook.active

    # Define a thin border style
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    # Define a date style (if needed)
    date_style = NamedStyle(name="date_style", number_format="DD.MM.YYYY")

    # Start filling data from the second row (assuming the template already has a header)
    start_row = 2
    for row_idx, row in enumerate(data, start=start_row):
        function(sheet, row, row_idx, date_style)

        # Apply borders to all cells in the current row
        for col_idx in range(1, number_of_columns):  # Adjust the range based on your columns
            cell = sheet.cell(row=row_idx, column=col_idx)
            cell.border = thin_border

    # Save the file to a BytesIO object to send it as an HTTP response
    file_stream = BytesIO()
    workbook.save(file_stream)
    file_stream.seek(0)

    # Prepare the HttpResponse with the filled template
    response = HttpResponse(
        file_stream.read(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )

    # Set the dynamic filename in the response header
    response['Content-Disposition'] = f'attachment; filename="{file_name}"'

    return response


def getFormsExcelFileName(filters):
    """
    Generate a filename for the Excel file based on the provided filters in the request.

    Args:
        request: Django request object containing GET parameters.

    Returns:
        str: A sanitized filename based on the filters.
    """
    # Get filters from the request with default fallback values
    start_date = filters.get('startDate', '')  # Use 'all' if no date is specified
    end_date = filters.get('endDate', '')  # Use 'all' if no activity type is specified
    organisation = filters.get('organisationFilter', '')  # Use 'all' if no activity type is specified
    if organisation != '':
        organisation = Organisation.objects.get(pk=organisation).name
    volunteer = filters.get('volunteerFilter', '')  # Use 'all' if no activity type is specified
    if volunteer != '':
        volunteer_obj = Volunteer.objects.get(pk=volunteer)
        volunteer = f"{volunteer_obj.user.first_name}_{volunteer_obj.user.last_name}"
    activity_type = filters.get('activityTypeFilter', '')  # Use 'all' if no activity type is specified

    start_date = sanitize_filename(start_date)
    end_date = sanitize_filename(end_date)
    organisation = sanitize_filename(organisation)
    volunteer = sanitize_filename(volunteer)
    activity_type = sanitize_filename(activity_type)

    # Construct the filename
    filename = f"Forme_{start_date}_{end_date}"
    if organisation:
        filename = filename + f"_{organisation}"
    if volunteer:
        filename = filename + f"_{volunteer}"
    if activity_type:
        filename = filename + f"_{activity_type}"
    return replace_special_characters(filename) + ".xlsx"


 # Sanitize filter values to make them safe for filenames
def sanitize_filename(value):
    return value.replace(" ", "_").replace("/", "-").replace("\\", "-")


def replace_special_characters(filename):
    replacements = {
        'č': 'c',
        'ć': 'c',
        'š': 's',
        'ž': 'z',
        'đ': 'dj',
    }
    for char, replacement in replacements.items():
        filename = filename.replace(char, replacement)

    return filename


def fill_rows_in_forms_excel(sheet, row, row_idx, date_style):
    volunteer_name = f"{row.volunteer.user.first_name} {row.volunteer.user.last_name}"
    organisation_name = row.volunteer.volunteer_organisation.first().name
    duration = calculate_total_time_duration(row.duration)
    duration_minutes = duration if duration else 0
    places = "; ".join([place.name for place in row.place.all()]) if row.place.all() else ""
    activities = "; ".join([activity.name for activity in row.activities.all()]) if row.activities.all() else ""
    description = row.description if row.description is not None else ""

    # Populate the template rows with data
    sheet.cell(row=row_idx, column=1).value = volunteer_name
    sheet.cell(row=row_idx, column=2).value = organisation_name
    sheet.cell(row=row_idx, column=3).value = row.date
    sheet.cell(row=row_idx, column=3).style = date_style
    sheet.cell(row=row_idx, column=4).value = duration_minutes
    sheet.cell(row=row_idx, column=5).value = row.get_activity_type_display()
    sheet.cell(row=row_idx, column=6).value = row.get_evaluation_display()
    sheet.cell(row=row_idx, column=7).value = places
    sheet.cell(row=row_idx, column=8).value = activities
    sheet.cell(row=row_idx, column=9).value = description



def calculate_total_time_duration(number):
    if not number:
        return "0h 0min"

    number = round(float(number), 2)
    hours = int(number)
    minutes = (number - hours) * 60

    return f"{hours}h {round(minutes)}min"


def fill_rows_in_volunteers_excel(sheet, row, row_idx, date_style):
    volunteer_name = f"{row.user.first_name}"
    volunteer_last_name = f"{row.user.last_name}"
    organisation_name = row.volunteer_organisation.first().name
    email = row.user.email
    phone_number = row.phone_number
    birth_date = row.birth_date
    age = calculate_age(birth_date)
    child = str(row.child) if hasattr(row, 'child') and row.child is not None else ""

    # Populate the template rows with data
    sheet.cell(row=row_idx, column=1).value = volunteer_name
    sheet.cell(row=row_idx, column=2).value = volunteer_last_name
    sheet.cell(row=row_idx, column=3).value = organisation_name
    sheet.cell(row=row_idx, column=4).value = email
    sheet.cell(row=row_idx, column=5).value = phone_number
    sheet.cell(row=row_idx, column=6).value = birth_date
    sheet.cell(row=row_idx, column=6).style = date_style
    sheet.cell(row=row_idx, column=7).value = age
    sheet.cell(row=row_idx, column=8).value = row.get_gender_display()
    sheet.cell(row=row_idx, column=9).value = child
    sheet.cell(row=row_idx, column=10).value = "Aktivan/na" if row.status else "Nije aktivan/na"
    sheet.cell(row=row_idx, column=11).value = "Posjeduje" if row.good_conduct_certificate else "Ne posjeduje"
    sheet.cell(row=row_idx, column=12).value = row.get_education_level_display()
    sheet.cell(row=row_idx, column=13).value = row.get_faculty_department_display() if row.faculty_department is not None else row.faculty_other_department
    sheet.cell(row=row_idx, column=14).value = row.get_employment_status_display()



def calculate_age(birth_date):
    today = datetime.today().date()
    age = today.year - birth_date.year #- ((today.month, today.day) < (birth_date.month, birth_date.day))
    return age



def getVolunteersExcelFileName(filters):
    """
    Generate a filename for the Excel file based on the provided filters in the request.

    Args:
        request: Django request object containing GET parameters.

    Returns:
        str: A sanitized filename based on the filters.
    """
    organisation = filters.get('organisationFilter', '')  # Use 'all' if no activity type is specified
    if organisation != '':
        organisation = Organisation.objects.get(pk=organisation).name

    organisation = sanitize_filename(organisation)

    filename = "Volonteri"
    if organisation:
        filename = filename + f"_{organisation}"

    return replace_special_characters(filename) + ".xlsx"

